import openpyxl

from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By

import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions

wb = openpyxl.load_workbook('../data/orange.xlsx')
sheet = wb['login']

un = sheet.cell(2,1).value #username
pwd = sheet.cell(2,2).value #password
exp_url = sheet.cell(2,3).value #expected url

driver = Chrome()
driver.implicitly_wait(10)
driver.get('https://opensource-demo.orangehrmlive.com')
driver.find_element(By.NAME,'username').send_keys(un)
driver.find_element(By.NAME, 'password').send_keys(pwd)

driver.find_element(By.XPATH,"//button[@type='submit']").click()

wait = WebDriverWait(driver, 10)

try:
    wait.until(expected_conditions.url_contains(exp_url))
    print('Successful login')
    status = 'PASS'
except:
    print('Login Fail')
    status = 'FAIL'

act_url = driver.current_url
sheet.cell(2,4).value = act_url
sheet.cell(2,5).value = status

# save the excel file and close it
wb.save('../data/orange.xlsx')
wb.close()













